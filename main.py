import pandas as pd
import numpy as np
from sklearn.linear_model import LinearRegression
from sklearn.metrics import r2_score
import os
import sys
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import itertools

def analyze_absorbance_data(input_file, custom_time_frame=None, epsilon=0.001):
    """
    Analyzes absorbance data from an Excel file, performing linear regression
    on each measurement column and generating a results Excel file.
    
    Args:
        input_file (str): Path to the Excel file containing absorbance data.
        custom_time_frame (tuple, optional): User-defined time frame (start, end) in seconds.
        epsilon (float): Extinction coefficient for v0/epsilon calculation.
    """
    print(f"Reading data from: {input_file}")
    
    # Read the Excel file starting from row 10 (0-indexed, so we use 9)
    df = pd.read_excel(input_file, skiprows=9, engine='openpyxl', decimal=',')    
    # Print a sample of data to diagnose issues
    print("\nSample of data (first 3 rows, first 5 columns):")
    print(df.iloc[:3, :5])
    print(f"Data types of first 5 columns: {df.iloc[:, :5].dtypes}")
    
    # Extract the time points (column 2)
    time_points = df.iloc[:, 1].values
    time_points_reshaped = time_points.reshape(-1, 1)
    
    # Process blanks - identify columns containing "Blank" in their name
    blank_columns = [col for col in df.columns if 'Blank' in str(col)]
    
    if blank_columns:
        print(f"\nFound {len(blank_columns)} blank columns: {blank_columns}")
        
        # Create a new DataFrame with blank-subtracted values
        df_processed = pd.DataFrame()
        df_processed[df.columns[0]] = df.iloc[:, 0]  # Copy first column (usually index)
        df_processed[df.columns[1]] = df.iloc[:, 1]  # Copy second column (time points)
        
        # Calculate mean blank value for each timepoint (row)
        blank_values = df[blank_columns].mean(axis=1)
        print("Calculated blank values (first 3 rows):")
        print(blank_values.head(3))
        
        # Subtract blank values from all other measurements
        for col in df.columns[2:]:  # Start from 3rd column (measurements)
            if col not in blank_columns:  # Skip blank columns
                # Convert both column and blank values to numeric before subtraction
                df_processed[col] = pd.to_numeric(df[col], errors='coerce') - blank_values
                
        print("Processed data with blank subtraction (first 3 rows, first 5 columns):")
        print(df_processed.iloc[:3, :5])
    else:
        print("No blank columns found - proceeding with original data.")
        df_processed = df
        
    # Get measurement columns (excluding blank columns)
    measurement_columns = [col for col in df_processed.columns[2:] if 'Blank' not in str(col)]
    
    # Ask user about time frame preference first
    if custom_time_frame is None:
        time_frame_choice = input("\nDo you want to: \n1. Define your own time frame\n2. Calculate the best time frame\nEnter 1 or 2: ").strip()
        
        if time_frame_choice == "1":
            # User wants to define their own time frame
            print("\nPlease enter your desired time frame (minimum 100 seconds):")
            start_time = float(input("Start time (seconds): "))
            end_time = float(input("End time (seconds): "))
            
            # Validate the user input
            if end_time - start_time < 100:
                print("Error: Time frame must be at least 100 seconds. Using default time frame.")
                # Set a default time frame using min/max values from the data
                min_time = np.min(time_points)
                max_time = np.max(time_points)
                custom_time_frame = (min_time, min_time + 100)
            else:
                custom_time_frame = (start_time, end_time)
        else:
            # User wants to calculate the best time frame
            optimal_time_frame, best_common_frame = find_optimal_time_frames(df_processed, time_points, measurement_columns)
            
            # Ask user if they want to use the suggested time frame or provide their own
            use_suggested = input(f"\nSuggested optimal common time frame: {best_common_frame[0]:.1f} - {best_common_frame[1]:.1f} seconds\nDo you want to use this time frame? (y/n): ").strip().lower()
            
            if use_suggested != 'y':
                print("\nPlease enter your desired time frame (minimum 100 seconds):")
                start_time = float(input("Start time (seconds): "))
                end_time = float(input("End time (seconds): "))
                
                # Validate the user input
                if end_time - start_time < 100:
                    print("Error: Time frame must be at least 100 seconds. Using suggested time frame instead.")
                    custom_time_frame = best_common_frame
                else:
                    custom_time_frame = (start_time, end_time)
            else:
                custom_time_frame = best_common_frame
    
    # Ask for epsilon value (or use default)
    use_default_epsilon = input(f"\nDefault extinction coefficient (epsilon) is set to {epsilon}.\nDo you want to use this value? (y/n): ").strip().lower()
    
    if use_default_epsilon != 'y':
        try:
            epsilon = float(input("Enter extinction coefficient (epsilon): "))
            print(f"Using epsilon value: {epsilon}")
        except ValueError:
            print(f"Invalid input. Using default epsilon value: {epsilon}")
    
    # Print the selected time frame
    print(f"\nAnalyzing data using time frame: {custom_time_frame[0]:.1f} - {custom_time_frame[1]:.1f} seconds")
    
    # Filter data based on the selected time frame
    time_mask = (time_points >= custom_time_frame[0]) & (time_points <= custom_time_frame[1])
    filtered_time_points = time_points[time_mask].reshape(-1, 1)
    
    # Dictionary to store results
    results = {
        'Condition': [],
        'Slope (k)': [],
        'v0 (abs k)': [],
        'v0/ε': [],  # New field for v0 divided by epsilon
        'Intercept (d)': [],
        'R-squared': [],
        'Time Frame': []
    }
    
    # Process each measurement column (excluding blank columns)
    for col in measurement_columns:
        try:
            # Get the column data
            absorbance = df_processed[col].values
            
            # Only skip if all values are NaN (the column is completely empty)
            if pd.isna(absorbance).all():
                print(f"Skipping column '{col}' because it's completely empty.")
                continue
            
            # Filter absorbance data based on time mask
            filtered_absorbance = absorbance[time_mask]
            
            # Skip if there are less than 2 valid data points (can't do regression)
            if np.sum(~pd.isna(filtered_absorbance)) < 2:
                print(f"Skipping column '{col}' because it has fewer than 2 valid data points in the selected time frame.")
                continue
                
            # Create and fit linear regression model
            model = LinearRegression()
            
            # Drop any rows where either time or absorbance is NaN
            valid_mask = ~pd.isna(filtered_absorbance)
            valid_time_points = filtered_time_points[valid_mask]
            valid_absorbance = filtered_absorbance[valid_mask]
            
            model.fit(valid_time_points, valid_absorbance)
            
            # Calculate predictions and R-squared
            y_pred = model.predict(valid_time_points)
            r_squared = r2_score(valid_absorbance, y_pred)
            
            # Store results
            slope = model.coef_[0]
            v0 = abs(slope)  # Calculate absolute value of slope
            v0_epsilon = v0 / epsilon  # Calculate v0/epsilon
            
            results['Condition'].append(col)
            results['Slope (k)'].append(slope)
            results['v0 (abs k)'].append(v0)
            results['v0/ε'].append(v0_epsilon)  # Add new calculated value
            results['Intercept (d)'].append(model.intercept_)
            results['R-squared'].append(r_squared)
            results['Time Frame'].append(f"{custom_time_frame[0]:.1f} - {custom_time_frame[1]:.1f}")
            
            print(f"Analyzed condition: {col}")
            print(f"  Slope (k): {slope:.6f}")
            print(f"  v0 (abs k): {v0:.6f}")
            print(f"  v0/ε: {v0_epsilon:.6f}")  # Print new calculated value
            print(f"  Intercept (d): {model.intercept_:.6f}")
            print(f"  R-squared: {r_squared:.6f}")
            print(f"  Time Frame: {custom_time_frame[0]:.1f} - {custom_time_frame[1]:.1f}")
            
        except Exception as e:
            print(f"Error processing column '{col}': {str(e)}")
            # Print a sample of the problematic column
            try:
                print(f"Sample of problematic column (first 3 rows):")
                print(df_processed.iloc[:3, df_processed.columns.get_loc(col)])
                print(f"Data type: {df_processed[col].dtype}")
            except:
                pass
    
    # Check if we have any results
    if not results['Condition']:
        print("No valid columns found for analysis. Please check your Excel file format.")
        return
    
    # Create a results DataFrame
    results_df = pd.DataFrame(results)
    
    # Generate output file name
    output_file = os.path.splitext(input_file)[0] + "_results.xlsx"
    1
    # Save results to a new Excel file with formatting
    create_formatted_excel(results_df, output_file, input_file, custom_time_frame, epsilon)
    
    print(f"\nResults saved to: {output_file}")
    
    # Create plots
    create_plots(df_processed, results_df, time_points, input_file, custom_time_frame)

    # Summary plot (new feature)  
    create_summary_plot(df_processed, results_df, time_points, input_file, custom_time_frame)  

def find_optimal_time_frames(df, time_points, measurement_columns):
    """
    Find the optimal time frame for each measurement column that gives the best linearity.
    
    Args:
        df (DataFrame): The data DataFrame.
        time_points (ndarray): Array of time points.
        measurement_columns (list): List of measurement column names to analyze (excluding blanks).
    
    Returns:
        dict: Dictionary of optimal time frames for each column.
        tuple: Best common time frame for all measurements.
    """
    print("\nAnalyzing optimal time frames for each measurement...")
    
    # Minimum window size in seconds
    min_window_size = 100
    
    # Dictionary to store optimal time frames and their R-squared values
    optimal_frames = {}
    
    # Process each measurement column
    for col in measurement_columns:
        absorbance = df[col].values
        
        # Skip completely empty columns
        if pd.isna(absorbance).all():
            continue
            
        print(f"Finding optimal time frame for: {col}")
        
        best_r2 = -1
        best_frame = None
        
        # Get minimum and maximum time points that have valid absorbance readings
        valid_mask = ~pd.isna(absorbance)
        if sum(valid_mask) < 2:
            continue
            
        valid_times = time_points[valid_mask]
        
        # Try different time windows
        # Start with the entire range and then try smaller windows
        min_time = np.min(valid_times)
        max_time = np.max(valid_times)
        
        # Get all possible time windows with at least min_window_size seconds
        possible_windows = []
        
        # Generate windows of different sizes
        for i, start_time in enumerate(valid_times):
            for j, end_time in enumerate(valid_times[i:], i):
                if end_time - start_time >= min_window_size:
                    possible_windows.append((start_time, end_time))
                    
        # Sort windows by size (smallest first) to prioritize smaller windows with similar R²
        possible_windows.sort(key=lambda x: x[1] - x[0])
        
        # Test each window
        for start_time, end_time in possible_windows:
            # Create time mask
            time_mask = (time_points >= start_time) & (time_points <= end_time)
            
            # Get filtered time points and absorbance
            filtered_time_points = time_points[time_mask].reshape(-1, 1)
            filtered_absorbance = absorbance[time_mask]
            
            # Skip if too few valid points
            valid_mask = ~pd.isna(filtered_absorbance)
            if sum(valid_mask) < 2:
                continue
                
            valid_time_points = filtered_time_points[valid_mask]
            valid_absorbance = filtered_absorbance[valid_mask]
            
            # Create and fit linear regression model
            try:
                model = LinearRegression()
                model.fit(valid_time_points, valid_absorbance)
                
                # Calculate R-squared
                y_pred = model.predict(valid_time_points)
                r_squared = r2_score(valid_absorbance, y_pred)
                
                # Check if this is the best window so far
                # If we already have an R² > 0.99, prefer the smaller window that's still > 0.98
                if r_squared > 0.99 and best_r2 > 0.98:
                    best_r2 = r_squared
                    best_frame = (start_time, end_time)
                    break  # We found an excellent window, we can stop here
                elif r_squared > best_r2:
                    best_r2 = r_squared
                    best_frame = (start_time, end_time)
            except Exception as e:
                continue
        
        # Store the best frame for this condition
        if best_frame is not None:
            optimal_frames[col] = {
                'frame': best_frame,
                'r2': best_r2
            }
            print(f"  Best time frame: {best_frame[0]:.1f} - {best_frame[1]:.1f} seconds (R² = {best_r2:.4f})")
    
    # Find the best common time frame
    best_common_frame = find_best_common_time_frame(optimal_frames, min_window_size)
    
    return optimal_frames, best_common_frame

def find_best_common_time_frame(optimal_frames, min_window_size):
    """
    Find the best common time frame that works well for all measurements.
    
    Args:
        optimal_frames (dict): Dictionary of optimal time frames for each condition.
        min_window_size (float): Minimum window size in seconds.
    
    Returns:
        tuple: Best common time frame (start, end).
    """
    if not optimal_frames:
        return (0, min_window_size)
    
    # Extract all optimal frame boundaries
    all_frames = [frame['frame'] for frame in optimal_frames.values()]
    
    # Find the latest start time and earliest end time
    latest_start = max(frame[0] for frame in all_frames)
    earliest_end = min(frame[1] for frame in all_frames)
    
    # Check if this window is valid (at least min_window_size seconds)
    if earliest_end - latest_start >= min_window_size:
        common_frame = (latest_start, earliest_end)
    else:
        # If no valid common window, find the best compromise
        # Collect all individual optimal frames
        all_start_times = [frame[0] for frame in all_frames]
        all_end_times = [frame[1] for frame in all_frames]
        
        # Calculate the median start and end times
        median_start = np.median(all_start_times)
        median_end = np.median(all_end_times)
        
        # Ensure the window is at least min_window_size seconds
        if median_end - median_start < min_window_size:
            # Extend the window to make it min_window_size seconds
            median_end = median_start + min_window_size
            
        common_frame = (median_start, median_end)
    
    # Test how well this common frame works for each condition
    print("\nEvaluating common time frame quality:")
    print(f"Proposed common time frame: {common_frame[0]:.1f} - {common_frame[1]:.1f} seconds")
    
    for condition, data in optimal_frames.items():
        original_frame = data['frame']
        original_r2 = data['r2']
        
        overlap_start = max(common_frame[0], original_frame[0])
        overlap_end = min(common_frame[1], original_frame[1])
        overlap_ratio = (overlap_end - overlap_start) / (original_frame[1] - original_frame[0])
        
        print(f"  {condition}: Overlap with optimal frame: {overlap_ratio:.2%}")
    
    return common_frame

def create_formatted_excel(results_df, output_file, input_file, time_frame, epsilon):
    """
    Creates a nicely formatted Excel file with the regression results.
    
    Args:
        results_df (DataFrame): DataFrame containing regression results.
        output_file (str): Path for the output Excel file.
        input_file (str): Path of the original input file for reference.
        time_frame (tuple): The time frame used for analysis.
        epsilon (float): Extinction coefficient used for v0/epsilon calculation.
    """
    # Create a new workbook and select active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Regression Results"
    
    # Add title and metadata
    ws['A1'] = "Absorbance Measurement Analysis"
    ws['A1'].font = Font(bold=True, size=14)
    
    ws['A2'] = f"Source file: {os.path.basename(input_file)}"
    ws['A3'] = f"Analysis date: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A4'] = f"Time frame used: {time_frame[0]:.1f} - {time_frame[1]:.1f} seconds"
    ws['A5'] = f"Extinction coefficient (ε): {epsilon}"
    
    # Add headers at row 7
    headers = ['Condition', 'Slope (k)', 'v0 (abs k)', 'v0/ε', 'Intercept (d)', 'R-squared', 'Time Frame']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        
        # Add border
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        cell.border = thin_border
    
    # Add data starting from row 8
    for row_idx, row in enumerate(results_df.itertuples(index=False), 8):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal='center')
            
            # Format numbers
            if col_idx in [2, 3, 4, 5]:  # Slope, v0, v0/epsilon, and intercept
                cell.number_format = '0.000000'
            elif col_idx == 6:  # R-squared
                cell.number_format = '0.000000'
                
                # Conditional formatting for R-squared
                if value > 0.95:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif value > 0.90:
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            # Add border
            thin_border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            cell.border = thin_border
    
    # Auto-adjust column widths
    for col_idx in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = 15
    
    # Add a summary section
    summary_row = len(results_df) + 10
    ws.cell(row=summary_row, column=1).value = "Summary"
    ws.cell(row=summary_row, column=1).font = Font(bold=True)
    
    ws.cell(row=summary_row + 1, column=1).value = "Total conditions analyzed:"
    ws.cell(row=summary_row + 1, column=2).value = len(results_df)
    
    # Only add these if we have results
    if len(results_df) > 0:
        ws.cell(row=summary_row + 2, column=1).value = "Average R-squared:"
        ws.cell(row=summary_row + 2, column=2).value = results_df['R-squared'].mean()
        ws.cell(row=summary_row + 2, column=2).number_format = '0.000000'
        
        ws.cell(row=summary_row + 3, column=1).value = "Average slope (k):"
        ws.cell(row=summary_row + 3, column=2).value = results_df['Slope (k)'].mean()
        ws.cell(row=summary_row + 3, column=2).number_format = '0.000000'
        
        ws.cell(row=summary_row + 4, column=1).value = "Average v0 (abs k):"
        ws.cell(row=summary_row + 4, column=2).value = results_df['v0 (abs k)'].mean()
        ws.cell(row=summary_row + 4, column=2).number_format = '0.000000'
        
        ws.cell(row=summary_row + 5, column=1).value = "Average v0/ε:"
        ws.cell(row=summary_row + 5, column=2).value = results_df['v0/ε'].mean()
        ws.cell(row=summary_row + 5, column=2).number_format = '0.000000'
        
        ws.cell(row=summary_row + 6, column=1).value = "Time frame used:"
        ws.cell(row=summary_row + 6, column=2).value = f"{time_frame[0]:.1f} - {time_frame[1]:.1f} seconds"
        
        ws.cell(row=summary_row + 7, column=1).value = "Extinction coefficient (ε):"
        ws.cell(row=summary_row + 7, column=2).value = epsilon
        ws.cell(row=summary_row + 7, column=2).number_format = '0.000000'
    
    # Save the workbook
    wb.save(output_file)

def create_plots(df, results_df, time_points, input_file, time_frame):
    """
    Creates plot visualizations of the absorbance data and regression lines.
    
    Args:
        df (DataFrame): Original data DataFrame.
        results_df (DataFrame): Results DataFrame with regression parameters.
        time_points (ndarray): Array of time points.
        input_file (str): Path of the input file for naming convention.
        time_frame (tuple): The time frame used for analysis.
    """
    # Check if we have any results to plot
    if len(results_df) == 0:
        print("No valid data to plot.")
        return
        
    # Create a folder for plots if it doesn't exist
    plots_folder = os.path.join(os.path.dirname(input_file), "plots_" + os.path.splitext(os.path.basename(input_file))[0])
    os.makedirs(plots_folder, exist_ok=True)
    
    # Create a plot for each condition
    for i, condition in enumerate(results_df['Condition']):
        plt.figure(figsize=(10, 6))
        
        try:
            # Get absorbance values
            absorbance = df[condition].values
            
            # Plot all data points (full range) with lower opacity
            valid_mask_all = ~pd.isna(absorbance)
            plt.scatter(time_points[valid_mask_all], absorbance[valid_mask_all], 
                        label='All data points', color='gray', alpha=0.3, s=20)
            
            # Filter for the selected time frame
            time_mask = (time_points >= time_frame[0]) & (time_points <= time_frame[1])
            filtered_time_points = time_points[time_mask]
            filtered_absorbance = absorbance[time_mask]
            
            # Filter out NaN values
            valid_mask = ~pd.isna(filtered_absorbance)
            valid_time_points = filtered_time_points[valid_mask].reshape(-1, 1)
            valid_absorbance = filtered_absorbance[valid_mask]
            
            # Get regression parameters
            slope = results_df.loc[i, 'Slope (k)']
            v0 = results_df.loc[i, 'v0 (abs k)']
            v0_epsilon = results_df.loc[i, 'v0/ε']
            intercept = results_df.loc[i, 'Intercept (d)']
            r2 = results_df.loc[i, 'R-squared']
            
            # Plot selected data points with higher opacity
            plt.scatter(valid_time_points, valid_absorbance, 
                        label=f'Selected data points ({time_frame[0]:.1f} - {time_frame[1]:.1f}s)', 
                        color='blue', alpha=0.7, s=30)
            
            # Plot regression line
            x_range = np.linspace(time_frame[0], time_frame[1], 100)
            y_line = slope * x_range + intercept
            plt.plot(x_range, y_line, 'r-', label=f'Regression line: y = {slope:.4f}x + {intercept:.4f}')
            
            # Mark the time frame with vertical lines
            plt.axvline(x=time_frame[0], color='green', linestyle='--', alpha=0.5, label='Time frame boundaries')
            plt.axvline(x=time_frame[1], color='green', linestyle='--', alpha=0.5)
            
            # Add labels and title including v0 and v0/epsilon
            plt.xlabel('Time (seconds)')
            plt.ylabel('Absorbance')
            plt.title(f'Absorbance vs Time for {condition}\nR² = {r2:.4f}, v0 = {v0:.4f}, v0/ε = {v0_epsilon:.4f}')
            plt.legend()
            plt.grid(True, alpha=0.3)
            
            # Save the plot
            safe_condition = condition.replace('/', '_').replace(':', '_').replace('\\', '_')
            safe_condition = ''.join(c if c.isalnum() or c in ['_', '-'] else '_' for c in safe_condition)
            plot_file = os.path.join(plots_folder, f"{safe_condition}.png")
            plt.savefig(plot_file)
            plt.close()
            
            print(f"Saved plot for {condition}")
            
        except Exception as e:
            print(f"Error creating plot for {condition}: {str(e)}")
            plt.close()


def create_summary_plot(df, results_df, time_points, input_file, time_frame):
    """
    Creates a summary plot showing all valid (filtered) data sets on one graph.
    """
    if len(results_df) == 0:
        print("No valid data for summary plot.")
        return
    plots_folder = os.path.join(os.path.dirname(input_file),  
                            "plots_" + os.path.splitext(os.path.basename(input_file))[0])  
    os.makedirs(plots_folder, exist_ok=True)  

    plt.figure(figsize=(10, 6))  
    # Use a colormap to differentiate lines  
    colors = plt.cm.tab20(np.linspace(0, 1, len(results_df)))  

    for i, condition in enumerate(results_df['Condition']):  
        try:  
            absorbance = df[condition].values  
            selected_mask = (time_points >= time_frame[0]) & (time_points <= time_frame[1])  
            filtered_time_points = time_points[selected_mask]  
            filtered_absorbance = absorbance[selected_mask]  

            valid_mask = ~pd.isna(filtered_absorbance)  
            valid_time_points = filtered_time_points[valid_mask]  
            valid_absorbance = filtered_absorbance[valid_mask]  

            if len(valid_time_points) < 2:  
                continue  

            plt.plot(valid_time_points, valid_absorbance, label=condition, color=colors[i])  
        except Exception as e:  
            print(f"Error processing summary plot for {condition}: {str(e)}")  

    plt.xlabel("Time (seconds)")  
    plt.ylabel("Absorbance")  
    plt.title("Summary Plot - All Conditions")  
    plt.legend()  
    plt.grid(True, alpha=0.3)  

    summary_plot_file = os.path.join(plots_folder, "Summary.png")  
    plt.savefig(summary_plot_file)  
    plt.close()  
    print(f"Saved summary plot with all conditions as {summary_plot_file}")  

def main():
    # Check if file is provided as command line argument
    if len(sys.argv) > 1:
        input_file = sys.argv[1]
    else:
        input_file = "20250605_kinetik_500_1000_1250_hcpo_PichiaE4_30nM.xlsx"
        # # Ask for file path
        # input_file = input("Please enter the path to the Excel file: ").strip()
        
        # # Handle quoted paths
        # if (input_file.startswith('"') and input_file.endswith('"')) or \
        #    (input_file.startswith("'") and input_file.endswith("'")):
        #     input_file = input_file[1:-1]
    
    # Check if file exists
    if not os.path.isfile(input_file):
        print(f"Error: File '{input_file}' not found.")
        sys.exit(1)
        
    # Process the file
    analyze_absorbance_data(input_file)

if __name__ == "__main__":
    main()
